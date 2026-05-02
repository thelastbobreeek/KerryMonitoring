import re
import sys
from pathlib import Path

import xlrd

CONFIG_FILE = Path("config.py")


def _clean_xls_cell(cell: xlrd.sheet.Cell) -> str:
    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return ""
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        value = cell.value
        return str(int(value)) if value == int(value) else str(value)
    return str(cell.value).strip()


def _clean_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _parse_rows(rows: list[tuple], limit: int | None = None) -> dict[str, dict]:
    result: dict[str, dict] = {}
    for i, row in enumerate(rows):
        if limit and i >= limit:
            break
        if len(row) < 5:
            continue

        our_brand, our_article, comp_brand, competitor_article, name = (
            _clean_value(row[0]),
            _clean_value(row[1]),
            _clean_value(row[2]),
            _clean_value(row[3]),
            _clean_value(row[4]),
        )

        if not our_article or not competitor_article or our_article == competitor_article:
            continue

        if our_article not in result:
            result[our_article] = {"brand": our_brand, "name": name, "competitors": {}}

        if competitor_article not in result[our_article]["competitors"]:
            result[our_article]["competitors"][competitor_article] = comp_brand

    return result


def read_articles_from_xls(path: str, limit: int | None = None) -> dict[str, dict]:
    if Path(path).suffix.lower() == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        return _parse_rows(rows, limit)

    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    total_rows = min(sheet.nrows, limit) if limit else sheet.nrows
    rows = [
        tuple(sheet.cell(row_idx, col) for col in range(sheet.ncols))
        for row_idx in range(1, total_rows)
    ]
    # xlrd cells need special cleaning — wrap them so _parse_rows can use _clean_value
    raw_rows = [
        tuple(_clean_xls_cell(sheet.cell(row_idx, col)) for col in range(sheet.ncols))
        for row_idx in range(1, total_rows)
    ]
    return _parse_rows(raw_rows, limit)


def _find_articles_block(content: str) -> tuple[int, int] | None:
    match = re.search(r"ARTICLES\s*=\s*\{", content)
    if not match:
        return None

    brace_start = match.end() - 1
    depth = 0
    for i, ch in enumerate(content[brace_start:]):
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                return match.start(), brace_start + i + 1
    return None


def _format_articles(articles: dict) -> str:
    lines = ["ARTICLES = {"]
    for our_article, data in articles.items():
        lines.append(f"    {repr(our_article)}: {{")
        lines.append(f'        "brand": {repr(data["brand"])},')
        lines.append(f'        "name": {repr(data["name"])},')
        lines.append(f'        "competitors": {{')
        for comp_article, comp_brand in data["competitors"].items():
            lines.append(f"            {repr(comp_article)}: {repr(comp_brand)},")
        lines.append("        },")
        lines.append("    },")
    lines.append("}")
    return "\n".join(lines)


def update_config_articles(articles: dict) -> None:
    content = CONFIG_FILE.read_text(encoding="utf-8")

    block_range = _find_articles_block(content)
    if block_range is None:
        print("Блок ARTICLES не найден в config.py")
        sys.exit(1)

    start, end = block_range
    new_content = content[:start] + _format_articles(articles) + content[end:]
    CONFIG_FILE.write_text(new_content, encoding="utf-8")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("file", nargs="?", default="articles.xls")
    parser.add_argument("--limit", type=int, default=None, help="обработать только первые N строк")
    parser.add_argument("--dry-run", action="store_true", help="показать результат без записи в config.py")
    args = parser.parse_args()

    if not Path(args.file).exists():
        print(f"Файл не найден: {args.file}")
        sys.exit(1)

    articles = read_articles_from_xls(args.file, limit=args.limit)
    if not articles:
        print("Артикулы не найдены в файле")
        sys.exit(1)

    print(f"Найдено {len(articles)} артикулов:")
    for our_article, data in articles.items():
        comp_preview = ", ".join(list(data["competitors"].keys())[:5])
        suffix = f"... (+{len(data['competitors']) - 5})" if len(data["competitors"]) > 5 else ""
        print(f"  {our_article} ({data['brand']}) → {comp_preview}{suffix}")

    if args.dry_run:
        print("\n--dry-run: config.py не изменён")
    else:
        update_config_articles(articles)
        print("\nconfig.py обновлён")
