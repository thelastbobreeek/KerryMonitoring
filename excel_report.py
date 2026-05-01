from io import BytesIO

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

_COL_WIDTH = 25


def build_report(rows: list[dict], all_brands: list[str]) -> bytes:
    """
    rows: list of {
        "brand": str,
        "article": str,
        "name": str,
        "our_price": float | None,
        "competitors": {brand: {"price": float, "article": str, "catalog": str, "article_id": str} | None}
    }
    all_brands: ordered list of all unique competitor brands (determines column order)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Цены"

    fixed_headers = ["Наш бренд", "Артикул", "Наименование", "Наша цена", "Лучшая цена конкурента"]
    comp_headers = list(all_brands)
    all_headers = fixed_headers + comp_headers
    ws.append(all_headers)

    for col_idx in range(1, len(all_headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTH

    for row_data in rows:
        our_price = row_data["our_price"]
        comp_results = row_data["competitors"]

        comp_prices = [v["price"] for v in comp_results.values() if v and v.get("price") is not None]
        best_comp_price = min(comp_prices) if comp_prices else None

        ws.append([
            row_data["brand"],
            row_data["article"],
            row_data["name"],
            our_price,
            best_comp_price,
        ])
        row_num = ws.max_row

        for col_idx, brand in enumerate(all_brands, start=6):
            comp = comp_results.get(brand)
            if not comp or comp.get("price") is None:
                continue

            price = comp["price"]
            article = comp["article"]
            catalog = comp.get("catalog", "")
            article_id = comp.get("article_id", "")

            cell = ws.cell(row=row_num, column=col_idx)

            if article_id and catalog and article:
                # TODO: verify article_id (from FindCatalog) is the correct ID for this URL pattern
                url = f"https://autopiter.ru/goods/{article.lower()}/{catalog.lower()}/id{article_id}"
                display = f"{price:.0f}({article})"
                cell.value = f'=HYPERLINK("{url}","{display}")'
                cell.font = Font(color="0563C1", underline="single")
            else:
                cell.value = f"{price:.0f}({article})"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
